import asyncio
import csv
import datetime
import io
import json
import os
import re
import logging
from typing import List, Optional, Any
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import httpx
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

# Load environment variables
load_dotenv(override=True)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("recruitment_pipeline")

app = FastAPI(
    title="Recruitment Screening Pipeline API",
    description="API for processing LinkedIn candidate profile CSV uploads against Ideal Candidate Profiles and fetching cleaned dataset items via Apify."
)

# Enable CORS for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve the project directory as static files (CSS, JS, JSON, etc.)
_STATIC_DIR = os.path.dirname(os.path.abspath(__file__))


@app.get("/")
async def serve_index():
    """Serves the main frontend page at http://localhost:8000."""
    return FileResponse(os.path.join(_STATIC_DIR, "index.html"))

APIFY_API_TOKEN = os.getenv("APIFY_API_TOKEN")
SCRAPINGDOG_API_KEY = os.getenv("SCRAPINGDOG_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_ENDPOINT = "https://api.openai.com/v1/chat/completions"
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")

EVALUATIONS_DIR = os.path.join(os.path.dirname(__file__), "evaluations")
MASTER_EXCEL_PATH = os.path.join(EVALUATIONS_DIR, "master_evaluations.xlsx")


def save_to_master_excel(evaluations: List[Any], candidates: List[dict]):
    """Appends candidate evaluation records and emails into a persistent master Excel spreadsheet."""
    os.makedirs(EVALUATIONS_DIR, exist_ok=True)
    headers = [
        "Timestamp", "Candidate Name", "Dummy Email", "LinkedIn URL",
        "Score", "Status", "Justification", "Outreach Email Draft"
    ]

    if os.path.exists(MASTER_EXCEL_PATH):
        try:
            wb = load_workbook(MASTER_EXCEL_PATH)
            ws = wb.active
        except Exception as e:
            logger.error(f"Error loading existing master excel, creating fresh workbook: {e}")
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for i, eval_res in enumerate(evaluations):
        cand_dict = candidates[i] if i < len(candidates) and isinstance(candidates[i], dict) else {}
        # Real email from profile if available, otherwise generate a placeholder
        real_email = cand_dict.get("email") or cand_dict.get("emailAddress") or ""
        clean_name = "".join(c for c in eval_res.name.lower() if c.isalnum() or c == " ").strip().replace(" ", ".")
        dummy_email = real_email if real_email else f"{clean_name if clean_name else 'candidate'}@candidate-placeholder.com"

        ws.append([
            now_str,
            eval_res.name,
            dummy_email,
            eval_res.linkedinUrl or "",
            eval_res.score,
            eval_res.status,
            eval_res.justification,
            eval_res.outreach_email or ""
        ])


    try:
        wb.save(MASTER_EXCEL_PATH)
        logger.info(f"Successfully appended {len(evaluations)} records to master Excel at: {MASTER_EXCEL_PATH}")
    except Exception as e:
        logger.error(f"Failed to save master excel: {e}")


# ── Pydantic Models ────────────────────────────────────────────────────────────

class PipelineResponse(BaseModel):
    urls: List[str] = Field(..., description="List of cleaned, validated LinkedIn URLs")
    icp_description: str = Field(..., description="The received Ideal Candidate Profile description")
    invalid_rows: List[str] = Field(default_factory=list, description="List of rows that failed validation")
    icp_data: Optional[Any] = Field(default=None, description="Cleaned profile data for the Ideal Candidate")
    candidates_data: Optional[List[Any]] = Field(default=None, description="Cleaned candidate profile data matching relevant schema")
    raw_data: Optional[List[Any]] = Field(default=None, description="Raw unfiltered Apify response data (ICP first if applicable, then candidates)")


class EvaluationRequest(BaseModel):
    icp_data: Optional[Any] = Field(None, description="Cleaned ICP profile dict")
    candidates_data: List[Any] = Field(..., description="List of cleaned candidate profile dicts")
    threshold: int = Field(75, description="The score threshold for qualification status")


class CandidateEvaluation(BaseModel):
    name: str
    linkedinUrl: Optional[str] = None
    score: int
    justification: str
    outreach_email: Optional[str] = None
    status: str  # "Qualified" | "Rejected"


class EvaluationResponse(BaseModel):
    results: List[CandidateEvaluation]


# ── Profile Cleaning ───────────────────────────────────────────────────────────

SCORING_FIELDS = {
    'id', 'publicIdentifier', 'linkedinUrl',
    'firstName', 'lastName', 'headline',
    'location', 'about', 'topSkills',
    'connectionsCount', 'followerCount', 'verified',
    'currentPosition', 'experience', 'education',
    'skills', 'certifications', 'publications',
    'honorsAndAwards', 'projects', 'volunteering',
    'patents', 'receivedRecommendations', 'recommendations',
    'languages', 'profilePicture', 'photo', 'openToWork',
    'has_profile_picture', 'skill_details',
}

LOGO_KEYS = {'companyLogo', 'schoolLogo', 'issuedByLogo'}


def strip_logos(items: list) -> list:
    return [{k: v for k, v in item.items() if k not in LOGO_KEYS} for item in (items or []) if isinstance(item, dict)]


def analyze_candidate_trajectory(profile: dict) -> dict:
    experience = profile.get('experience') or []
    if not isinstance(experience, list):
        experience = []
    
    role_count = len(experience)
    flagged_gaps = []
    flagged_title_inflation = []
    general_flags = []
    
    total_months = 0
    durations_months = []
    
    senior_titles = {'director', 'head', 'vp', 'vice president', 'chief', 'cto', 'ceo', 'cpo', 'lead', 'principal'}
    
    for exp in experience:
        if not isinstance(exp, dict):
            continue
        title = (exp.get('title') or exp.get('position') or '').lower()
        duration_str = exp.get('duration') or ''
        
        y_match = re.search(r'(\d+)\s*yr', duration_str, re.IGNORECASE)
        m_match = re.search(r'(\d+)\s*mo', duration_str, re.IGNORECASE)
        
        y = int(y_match.group(1)) if y_match else 0
        m = int(m_match.group(1)) if m_match else 0
        months = y * 12 + m
        if months > 0:
            durations_months.append(months)
            total_months += months
            
        if any(st in title for st in senior_titles):
            if months > 0 and months < 8:
                role_name = exp.get('title') or exp.get('position')
                flagged_title_inflation.append(f"Senior role '{role_name}' held for under 8 months ({duration_str}).")
                
    avg_tenure_years = round((total_months / len(durations_months) / 12), 1) if durations_months else None
    total_exp_years = round(total_months / 12, 1) if total_months > 0 else None
    
    if avg_tenure_years and avg_tenure_years < 1.2 and role_count >= 3:
        general_flags.append(f"Short average tenure pattern: ~{avg_tenure_years} years per role across {role_count} roles.")
        
    about_text = profile.get('about') or ''
    if not about_text or len(str(about_text).strip()) < 40:
        general_flags.append("Sparse or templated About section (low self-positioning signal).")
        
    return {
        "total_experience_years": total_exp_years,
        "average_tenure_years": avg_tenure_years,
        "role_count": role_count,
        "flagged_gaps": flagged_gaps,
        "flagged_title_inflation": flagged_title_inflation,
        "general_flags": general_flags
    }


def clean_profile(profile: dict) -> dict:
    if not isinstance(profile, dict):
        return profile
    result = {}
    for field in SCORING_FIELDS:
        if field not in profile:
            continue
        val = profile[field]
        if field in ('currentPosition', 'experience', 'education', 'certifications', 'honorsAndAwards', 'projects', 'volunteering', 'patents'):
            result[field] = strip_logos(val)
        else:
            result[field] = val
            
    result['hasProfilePicture'] = bool(profile.get('profilePicture') or profile.get('photo') or profile.get('has_profile_picture'))
    result['trajectory_analysis'] = analyze_candidate_trajectory(result)
    return result


def clean_profiles(profiles: list) -> list:
    return [clean_profile(p) for p in profiles]


# ── Scrapingdog Scraping ───────────────────────────────────────────────────────

def extract_linkedin_id(url: str) -> str:
    url = url.strip().rstrip('/')
    if "/in/" in url:
        return url.split("/in/")[-1].split('?')[0]
    return url


def map_scrapingdog_to_apify(sd_profile: dict) -> dict:
    if not isinstance(sd_profile, dict):
        return sd_profile
    
    # Extract name parts
    full_name = sd_profile.get("full_name", "")
    first_name = sd_profile.get("first_name", "")
    last_name = sd_profile.get("last_name", "")
    if not first_name and not last_name and full_name:
        parts = full_name.split()
        if len(parts) > 1:
            first_name = parts[0]
            last_name = " ".join(parts[1:])
        else:
            first_name = full_name

    apify_profile = {
        "firstName": first_name,
        "lastName": last_name,
        "publicIdentifier": sd_profile.get("public_identifier", ""),
        "linkedinUrl": f"https://www.linkedin.com/in/{sd_profile.get('public_identifier', '')}/" if sd_profile.get("public_identifier") else "",
        "headline": sd_profile.get("headline", ""),
        "location": sd_profile.get("location", ""),
        "about": sd_profile.get("about", ""),
        "topSkills": sd_profile.get("top_skills", []),
        "connectionsCount": sd_profile.get("connections", 0),
        "followerCount": sd_profile.get("followers", 0),
        "verified": sd_profile.get("verified", False),
        "experience": [],
        "education": [],
        "skills": sd_profile.get("skills", []),
        "certifications": sd_profile.get("certifications", []),
        "languages": sd_profile.get("languages", []),
        "projects": sd_profile.get("projects", []),
    }
    
    # Map experience
    for exp in sd_profile.get("experience", []):
        apify_profile["experience"].append({
            "title": exp.get("position", ""),
            "companyName": exp.get("company_name", ""),
            "location": exp.get("location", ""),
            "description": exp.get("description", ""),
            "start": exp.get("starts_at", ""),
            "end": exp.get("ends_at", "")
        })
        
    # Map education
    for edu in sd_profile.get("education", []):
        apify_profile["education"].append({
            "schoolName": edu.get("school_name", ""),
            "degreeName": edu.get("degree_name", ""),
            "fieldOfStudy": edu.get("field_of_study", ""),
            "start": edu.get("starts_at", ""),
            "end": edu.get("ends_at", "")
        })
        
    return apify_profile


def map_anchor_to_apify(anchor_profile: dict) -> dict:
    if not isinstance(anchor_profile, dict):
        return anchor_profile
    
    # Extract public identifier
    public_id = anchor_profile.get("public_identifier") or ""
    if not public_id and anchor_profile.get("url"):
        public_id = extract_linkedin_id(anchor_profile["url"])
        
    full_name = anchor_profile.get("full_name") or ""
    first_name = anchor_profile.get("first_name") or ""
    last_name = anchor_profile.get("last_name") or ""
    if not first_name and not last_name and full_name:
        parts = full_name.split()
        if len(parts) > 1:
            first_name = parts[0]
            last_name = " ".join(parts[1:])
        else:
            first_name = full_name

    apify_profile = {
        "firstName": first_name,
        "lastName": last_name,
        "publicIdentifier": public_id,
        "linkedinUrl": anchor_profile.get("url") or (f"https://www.linkedin.com/in/{public_id}/" if public_id else ""),
        "headline": anchor_profile.get("headline") or "",
        "location": anchor_profile.get("city") or anchor_profile.get("country") or "",
        "about": anchor_profile.get("summary") or "",
        "connectionsCount": 0,
        "followerCount": anchor_profile.get("follower_count") or 0,
        "verified": anchor_profile.get("show_verification_badge") or False,
        "experience": [],
        "education": [],
        "skills": [],
        "certifications": anchor_profile.get("certifications") or [],
        "languages": anchor_profile.get("languages") or [],
        "projects": [],
        "openToWork": anchor_profile.get("open_to_work") or False,
        "profilePicture": anchor_profile.get("profile_pic_url") or "",
    }
    
    # Map experiences
    for exp in anchor_profile.get("experiences") or []:
        if not isinstance(exp, dict):
            continue
        starts_at = exp.get("starts_at") or ""
        ends_at = exp.get("ends_at") or "Present"
        duration = f"{starts_at} - {ends_at}" if starts_at else ""
        apify_profile["experience"].append({
            "title": exp.get("title") or "",
            "companyName": exp.get("company") or "",
            "description": exp.get("description") or "",
            "duration": duration,
            "startDate": {"text": starts_at} if starts_at else None,
            "endDate": {"text": ends_at} if ends_at else None,
        })
        
    # Map education
    for edu in anchor_profile.get("education") or []:
        if not isinstance(edu, dict):
            continue
        starts_at = edu.get("starts_at") or ""
        ends_at = edu.get("ends_at") or ""
        apify_profile["education"].append({
            "schoolName": edu.get("school") or "",
            "degreeName": edu.get("degree") or "",
            "fieldOfStudy": edu.get("field_of_study") or "",
            "startDate": {"text": starts_at} if starts_at else None,
            "endDate": {"text": ends_at} if ends_at else None,
        })
        
    # Map skills
    for skill in anchor_profile.get("skills") or []:
        if isinstance(skill, str):
            apify_profile["skills"].append({"name": skill})
        elif isinstance(skill, dict):
            apify_profile["skills"].append({"name": skill.get("name") or ""})
            
    return apify_profile


async def scrape_single_profile(url: str, client: httpx.AsyncClient) -> Optional[dict]:
    """
    Makes an individual synchronous Apify actor API call for one LinkedIn URL.
    Returns the raw candidate profile dictionary, or None on failure.
    """
    apify_token = os.getenv("APIFY_API_TOKEN")
    actor_id = os.getenv("APIFY_ACTOR_ID", "anchor~linkedin-profile-enrichment")
    
    endpoint = f"https://api.apify.com/v2/acts/{actor_id}/run-sync-get-dataset-items"
    params = {"token": apify_token}
    # anchor~linkedin-profile-enrichment requires startUrls as an array of {"url": "..."} objects
    payload = {"startUrls": [{"url": url}]}
    
    try:
        logger.info(f"  → Scraping via Apify ({actor_id}): {url}")
        response = await client.post(endpoint, params=params, json=payload, timeout=180.0)
        if response.status_code in (200, 201):
            data = response.json()
            profile = None
            if isinstance(data, list) and len(data) > 0:
                profile = data[0]
            elif isinstance(data, dict):
                profile = data
                
            if profile and "anchor~linkedin-profile-enrichment" in actor_id:
                profile = map_anchor_to_apify(profile)
            return profile
        else:
            logger.error(f"Apify HTTP {response.status_code} for {url}: {response.text[:200]}")
    except Exception as e:
        logger.error(f"Exception scraping {url} via Apify: {e}")
    return None




async def scrape_all_parallel(urls: List[str]) -> List[Optional[dict]]:
    """
    Makes N simultaneous individual Apify calls — one per URL — using asyncio.gather.
    Returns results in the SAME ORDER as the input urls list.
    """
    if not urls:
        return []
    logger.info(f"Launching {len(urls)} parallel Apify calls simultaneously...")
    async with httpx.AsyncClient(timeout=120.0) as client:
        tasks = [scrape_single_profile(url, client) for url in urls]
        results = await asyncio.gather(*tasks)
    logger.info(f"All {len(urls)} parallel calls complete. "
                f"Successful: {sum(1 for r in results if r is not None)}/{len(urls)}")
    return list(results)


# ── Groq LLM Evaluation ────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are an expert recruitment screening evaluator scoring candidate profiles against an Ideal Candidate Profile (ICP).

Evaluate the candidate across all 7 essential recruitment dimensions:
1. Identity & positioning (Headline clarity, seniority proxy, completeness)
2. Career trajectory & tenure (Progression, job stability vs job-hopping, company caliber)
3. Skills & technical depth (Evidence of skills inside experience descriptions vs raw tags)
4. Education & credentials (Degree, field, institution, honors/awards, certifications)
5. Evidence of impact (Projects built, publications, patents, recommendations vouching for them)
6. Network & social proof (Language alignment, professional presence)
7. Red / Yellow Flags (Check trajectory_analysis for unexplained gaps, title inflation relative to tenure, sparse about section)

You must respond with this exact JSON structure, nothing else:

{{
  "name": "candidate full name",
  "linkedinUrl": "candidate linkedin url",
  "score": <integer 1-100>,
  "justification": "3-4 sentences explaining the score. Must cite specific details from their profile: degree, tenure, tools, projects, and any flagged risks. Do not be vague.",
  "outreach_email": "personalized cold email under 150 words for this candidate. Must reference their specific accomplishments, projects, or experience by name."
}}

Scoring rubric (total 100 points):
- Education & Credentials Match (relative to ICP): 0-25
- Career Trajectory & Tenure Match (relative to ICP): 0-30
- Technical Depth & Demonstrated Impact (projects/pubs/code evidence): 0-30
- Overall Positioning, Network & Flag Assessment: 0-15

Rules:
- Never invent information not present in the profile.
- Explicitly factor in any flagged risks in trajectory_analysis.
- Be rigorous. A score above {threshold} means genuinely qualified and strongly aligned with the ICP."""


async def evaluate_single_candidate(
    icp_data: Optional[dict],
    candidate: dict,
    client: httpx.AsyncClient,
    threshold: int
) -> Optional[CandidateEvaluation]:
    """
    Sends one candidate + ICP context to OpenAI GPT for evaluation.
    Returns a CandidateEvaluation or None on failure.
    """
    openai_key = os.getenv("OPENAI_API_KEY")
    if not openai_key:
        logger.error("OPENAI_API_KEY not set — cannot evaluate candidates.")
        return None

    name = f"{candidate.get('firstName', '')} {candidate.get('lastName', '')}".strip() or "Unknown"
    linkedin_url = candidate.get("linkedinUrl", "")

    icp_block = json.dumps(icp_data, indent=2) if icp_data else "Not provided."
    candidate_block = json.dumps(candidate, indent=2)

    user_message = (
        f"IDEAL CANDIDATE PROFILE:\n{icp_block}\n\n"
        f"CANDIDATE TO EVALUATE:\n{candidate_block}\n\n"
        "Evaluate this candidate strictly against the ICP and scoring rubric. "
        "Respond with the JSON structure only."
    )

    formatted_system_prompt = SYSTEM_PROMPT.format(threshold=threshold)

    payload = {
        "model": OPENAI_MODEL,
        "temperature": 0.1,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": formatted_system_prompt},
            {"role": "user", "content": user_message},
        ],
    }

    headers = {
        "Authorization": f"Bearer {openai_key}",
        "Content-Type": "application/json",
    }

    try:
        logger.info(f"  → Evaluating via OpenAI ({OPENAI_MODEL}): {name} ({linkedin_url}) with threshold {threshold}")
        response = await client.post(OPENAI_ENDPOINT, json=payload, headers=headers, timeout=60.0)
        if response.status_code == 200:
            raw = response.json()
            content = raw["choices"][0]["message"]["content"]
            parsed = json.loads(content)
            score = int(parsed.get("score", 0))
            return CandidateEvaluation(
                name=parsed.get("name", name),
                linkedinUrl=parsed.get("linkedinUrl", linkedin_url),
                score=score,
                justification=parsed.get("justification", ""),
                outreach_email=parsed.get("outreach_email"),
                status="Qualified" if score >= threshold else "Rejected",
            )
        else:
            logger.error(f"OpenAI HTTP {response.status_code} for {name}: {response.text[:300]}")
    except Exception as e:
        logger.error(f"Exception evaluating {name}: {e}")
    return None


async def evaluate_all_parallel(
    icp_data: Optional[dict],
    candidates: List[dict],
    threshold: int
) -> List[Optional[CandidateEvaluation]]:
    """Fire one LLM request per candidate simultaneously."""
    if not candidates:
        return []
    logger.info(f"Launching {len(candidates)} parallel OpenAI ({OPENAI_MODEL}) evaluation calls with threshold {threshold}...")
    async with httpx.AsyncClient(timeout=90.0) as client:
        tasks = [evaluate_single_candidate(icp_data, c, client, threshold) for c in candidates]
        results = await asyncio.gather(*tasks)
    successful = sum(1 for r in results if r is not None)
    logger.info(f"All {len(candidates)} evaluations complete. Successful: {successful}/{len(candidates)}")
    return list(results)


# ── Endpoints ──────────────────────────────────────────────────────────────────

@app.post("/run-pipeline", response_model=PipelineResponse)
async def run_pipeline(
    csv_file: Optional[UploadFile] = File(None, description="CSV file containing LinkedIn profile URLs"),
    icp_description: Optional[str] = Form("", description="Ideal Candidate Profile URL"),
    fetch_profiles: bool = Form(True, description="Whether to trigger Apify profile scraping"),
    mock_json: Optional[str] = Form(None, description="Raw JSON payload for mock processing and filtering")
):
    """
    Accepts a CSV of LinkedIn URLs + an ICP URL, then fires N simultaneous Apify calls
    (one per URL). Results are returned in input order:
      results[0] = ICP profile
      results[1..5] = candidates in CSV order
    """
    if not icp_description or not icp_description.strip():
        raise HTTPException(status_code=400, detail="Ideal Candidate Profile (ICP) URL is mandatory. Testing without an ICP URL is not allowed.")

    validated_urls: List[str] = []
    invalid_rows: List[str] = []

    if csv_file and csv_file.filename:
        if not csv_file.filename.endswith('.csv'):
            logger.warning(f"File '{csv_file.filename}' does not have a .csv extension.")
        try:
            content_bytes = await csv_file.read()
            content_str = content_bytes.decode('utf-8', errors='ignore')
            csv_reader = csv.reader(io.StringIO(content_str))
            raw_rows = [row[0].strip() for row in csv_reader if row and row[0].strip()]

            if raw_rows:
                first_row_lower = raw_rows[0].lower()
                header_keywords = ["url", "linkedin", "profile", "link", "candidate", "href"]
                is_header = (
                    any(kw in first_row_lower for kw in header_keywords)
                    and "linkedin.com/in/" not in first_row_lower
                )
                start_index = 1 if is_header else 0
                for entry in raw_rows[start_index:]:
                    if "linkedin.com/in/" in entry.lower():
                        validated_urls.append(entry)
                    else:
                        logger.warning(f"Invalid URL skipped: '{entry}'")
                        invalid_rows.append(entry)
        except Exception as e:
            logger.error(f"Error reading CSV: {e}")

    candidates_data: List[Any] = []
    icp_data: Optional[Any] = None

    raw_data: Optional[List[Any]] = None

    # ── Mock JSON path (no Apify calls — filter and return directly) ──────────
    if mock_json and mock_json.strip():
        logger.info("Processing mock_json input...")
        try:
            parsed = json.loads(mock_json)
            # Preserve raw before cleaning
            if isinstance(parsed, list):
                raw_data = parsed
                is_icp_url = bool(icp_description and "linkedin.com/in/" in icp_description.lower())
                if is_icp_url and parsed:
                    icp_data = clean_profile(parsed[0])
                    candidates_data = clean_profiles(parsed[1:])
                else:
                    candidates_data = clean_profiles(parsed)
            elif isinstance(parsed, dict):
                raw_data = [parsed]
                if "icp_data" in parsed or "candidates_data" in parsed:
                    if parsed.get("icp_data"):
                        icp_data = clean_profile(parsed["icp_data"])
                    if parsed.get("candidates_data"):
                        candidates_data = clean_profiles(parsed["candidates_data"])
                else:
                    candidates_data = [clean_profile(parsed)]
        except Exception as e:
            logger.error(f"Failed to parse mock_json: {e}")

    # ── Live Apify path — N parallel individual calls ─────────────────────────
    elif fetch_profiles:
        is_icp_url = bool(icp_description and "linkedin.com/in/" in icp_description.lower())

        urls_to_fetch: List[str] = []
        if is_icp_url:
            urls_to_fetch.append(icp_description.strip())
        urls_to_fetch.extend(validated_urls)

        if urls_to_fetch:
            raw_results = await scrape_all_parallel(urls_to_fetch)  # Raw, unfiltered
            raw_data = [r for r in raw_results if r is not None]

            if is_icp_url:
                icp_data = clean_profile(raw_results[0]) if raw_results[0] else None
                candidates_data = [clean_profile(r) for r in raw_results[1:] if r is not None]
            else:
                candidates_data = [clean_profile(r) for r in raw_results if r is not None]

    return PipelineResponse(
        urls=validated_urls,
        icp_description=icp_description or "",
        invalid_rows=invalid_rows,
        icp_data=icp_data,
        candidates_data=candidates_data,
        raw_data=raw_data,
    )


@app.post("/evaluate-candidates", response_model=EvaluationResponse)
async def evaluate_candidates(request: EvaluationRequest):
    """
    Receives ICP + candidate profiles, fires one OpenAI LLM call per candidate
    simultaneously, returns structured scores/justifications/outreach emails,
    and logs evaluation records to the persistent master Excel file.
    """
    if not request.icp_data:
        raise HTTPException(status_code=400, detail="Ideal Candidate Profile (ICP) data is required for candidate evaluation.")

    openai_key = os.getenv("OPENAI_API_KEY")
    if not openai_key:
        raise HTTPException(
            status_code=503,
            detail="OPENAI_API_KEY is not configured. Add it to your .env file and restart the server."
        )

    candidates = request.candidates_data or []
    if not candidates:
        return EvaluationResponse(results=[])

    raw_results = await evaluate_all_parallel(request.icp_data, candidates, request.threshold)

    # Replace None failures with a placeholder so positions stay aligned
    results = []
    for i, r in enumerate(raw_results):
        if r is not None:
            results.append(r)
        else:
            cand = candidates[i] if i < len(candidates) else {}
            name = f"{cand.get('firstName', '')} {cand.get('lastName', '')}".strip() or f"Candidate {i+1}"
            results.append(CandidateEvaluation(
                name=name,
                linkedinUrl=cand.get("linkedinUrl"),
                score=0,
                justification="Evaluation failed — LLM call returned no response.",
                outreach_email=None,
                status="Rejected",
            ))

    # Append evaluations to master Excel file
    save_to_master_excel(results, candidates)

    return EvaluationResponse(results=results)


@app.get("/download-master-excel")
async def download_master_excel():
    """Returns the persistent master Excel file containing all evaluated candidates."""
    if not os.path.exists(MASTER_EXCEL_PATH):
        raise HTTPException(status_code=404, detail="Master Excel file does not exist yet. Run candidate evaluations first.")
    return FileResponse(
        path=MASTER_EXCEL_PATH,
        filename="master_evaluations.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# Mount static files LAST so API routes take priority.
# This serves styles.css, combined_mock.json, xlsx CDN fallback, etc.
app.mount("/", StaticFiles(directory=_STATIC_DIR, html=True), name="static")
